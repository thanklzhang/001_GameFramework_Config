# -*- coding: utf-8 -*-
#!/usr/bin/python

# import copy
# import getopt
import json
import os.path
# import re
import sys
import codecs

from jinja2 import Environment, FileSystemLoader

from openpyxl import load_workbook

# Get Command Line Arguments

enums = []

def main(argv):
   
    list_dirs = os.walk('data_table')
    for root,dirs,files in list_dirs:
        for f in files:
            splitStr = os.path.splitext(f)
            path_without_ext = splitStr[0]
            ext = splitStr[1]
            if ext == '.xlsx' and not '~$' in path_without_ext:
                input_path = os.path.join(root, f)

                cs_out_dictionary = "cs_define"
                json_out_dictionary = "json"
                #cs define files
                gen_cs_define_file(input_path,cs_out_dictionary)
                #cs json files
                gen_json_file(input_path,json_out_dictionary)
                


################## 生成单个cs 定义文件
def gen_cs_define_file(input_file,output_dictionary):
    data_list,class_name = get_table_head_list(input_file)
    env = Environment(loader=FileSystemLoader('./'))
    template = env.get_template('template/cs_template.j2')
    result = template.render( name=class_name,list=data_list)

    if not os.path.exists(output_dictionary):
        os.mkdir(output_dictionary)

    with codecs.open(f'{output_dictionary}/{class_name}.cs', "w", 'utf8') as f:
        f.write(result)
    print("finish generate cs : " + class_name + ".cs")

################## 生成单个json 文件
def gen_json_file(input_file,output_dictionary):
    json_data,class_name = get_table_data_list(input_file,output_dictionary)

    json_str = json.dumps(json_data,sort_keys=False, indent=4, separators=(',', ': '))
    
    with codecs.open(f'{output_dictionary}/{class_name}.json', "w", 'utf8') as f:
        f.write(json_str)

################## 获取表格的前几行表头数据
def get_table_head_list(input_file):
    # print(input_file)
    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    name_obj_list = ws[1]
    type_obj_list = ws[2]
    comment_obj_list = ws[3]

    head_list = []
    for col_index,name_obj in enumerate(name_obj_list):
        # #过滤转表符
        # if 0 == col_index:
        #     continue
        name_value = name_obj.value
        type_value = type_obj_list[col_index].value
        comment_value = comment_obj_list[col_index].value
        head_data = HeadDefine(name_value,type_value,comment_value)
        head_list.append(head_data)
        # print(head_data.name + " " + head_data.type + " " + head_data.comment)

    #print(input_file)
    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]

    return head_list,class_name
   

################## 获取表格的数据部分(json)
def get_table_data_list(input_file, output_dictionary):
    # print(input_file)
    head_list,class_name = get_table_head_list(input_file)

    wb = load_workbook(input_file)
    ws = wb.worksheets[0]
    json_data = []
    for row_index,row_data in enumerate(ws):
        curr_data = {}
        if row_index >= 3:
            is_gen_data = True
            for col_index,data_obj in enumerate(row_data):
                #转表符 不生成数据
                if 0 == col_index:
                    if data_obj.value != '#':
                        is_gen_data = False
                        break
                
                field_name = head_list[col_index].name
                field_type = head_list[col_index].type
                curr_data[field_name] = data_obj.value
            if is_gen_data:
                json_data.append(curr_data)

    file_name = os.path.split(input_file)[1]
    class_name = os.path.splitext(file_name)[0]
    return json_data,class_name
   


##################
class HeadDefine():
    def __init__(self,name,type,comment):
        self.name = name
        self.type = type
        self.comment = comment


##################
if __name__ == "__main__":
    main(sys.argv)
